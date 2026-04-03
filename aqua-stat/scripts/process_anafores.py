import os
import json
import pandas as pd
import numpy as np
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

BASE_DIR = os.path.join(PROJECT_ROOT, "data", "ANAFORES")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "stats")
YEARS = range(2020, 2026)

SHEET_NAMES = {
    "flow": "ΠΑΡΟΧΗ",
    "hours": "ΩΡΕΣ ΛΕΙΤΟΥΡΓΙΑΣ",
    "starts": "ΕΚΚΙΝΗΣΕΙΣ",
    "energy": "ΚΑΤΑΝΑΛΩΣΗ ΕΝΕΡΓΕΙΑΣ",
}

SHORT_CYCLING_THRESHOLD = 3.0
HIGH_UTIL_HOURS = 22

# Three dashboard cards: (1) external feed line, (2) borehole + A3.9, (3) distribution FIT 3.9 + A1/A2
LINE_CARD_LABELS = (
    "Είσοδος Άγιος Γεωργίος — FIT 1.9 (υδραγωγείο)",
    "Γεώτρηση — FIT 2.9 + αντλία A3.9",
    "Προς Δ.Δ. Κωστάκιων — FIT 3.9 + αντλίες A1.9 & A2.9",
)


def read_all_monthly_files():
    """Read all monthly XLS files and merge their 4 worksheets."""
    all_data = []

    for year in YEARS:
        folder = os.path.join(BASE_DIR, f"STATISTIKA_ETOYS_{year}")
        if not os.path.exists(folder):
            print(f"  Skipping (not found): STATISTIKA_ETOYS_{year}/")
            continue

        for month in range(1, 13):
            fname = f"MONTH_{month}_{year}.xls"
            fpath = os.path.join(folder, fname)
            if not os.path.exists(fpath):
                print(f"  Skipping (not found): {fname}")
                continue

            try:
                df_flow = pd.read_excel(fpath, sheet_name=SHEET_NAMES["flow"])
                df_hours = pd.read_excel(fpath, sheet_name=SHEET_NAMES["hours"])
                df_starts = pd.read_excel(fpath, sheet_name=SHEET_NAMES["starts"])
                df_energy = pd.read_excel(fpath, sheet_name=SHEET_NAMES["energy"])

                merged = pd.merge(df_flow, df_hours, on="HM/NIA", how="outer")
                merged = pd.merge(
                    merged, df_starts, on="HM/NIA", how="outer", suffixes=("", "_STARTS")
                )
                merged = pd.merge(merged, df_energy, on="HM/NIA", how="outer")

                all_data.append(merged)
                print(f"  OK: {fname} ({len(merged)} rows)")
            except Exception as e:
                print(f"  ERROR: {fname} - {e}")

    if not all_data:
        raise RuntimeError("No data files were successfully read. Check folder paths.")

    return pd.concat(all_data, ignore_index=True)


def clean_columns(df):
    """Map Greek headers to hydraulic / equipment semantics (FIT vs A*.9 pumps)."""

    def classify(col):
        c = str(col).upper()
        if "HM/NIA" in c:
            return "Date"
        if "FIT" in c and "1.9" in c:
            return "Flow_Agios_m3"
        if "FIT" in c and "2.9" in c:
            return "Flow_Borehole_m3"
        if "FIT" in c and "3.9" in c:
            return "Flow_Distribution_m3"
        if "A1.9" in c and "STARTS" in c:
            return "Starts_A1_9"
        if "A2.9" in c and "STARTS" in c:
            return "Starts_A2_9"
        if "A3.9" in c and "STARTS" in c:
            return "Starts_A3_9"
        if "A1.9" in c:
            return "Hours_A1_9"
        if "A2.9" in c:
            return "Hours_A2_9"
        if "A3.9" in c:
            return "Hours_A3_9"
        if "KWH" in c or "ΕΝΕΡΓ" in c:
            return "Energy_kWh"
        return c

    df.columns = [classify(c) for c in df.columns]

    expected = {
        "Date",
        "Flow_Agios_m3",
        "Flow_Borehole_m3",
        "Flow_Distribution_m3",
        "Hours_A1_9",
        "Hours_A2_9",
        "Hours_A3_9",
        "Starts_A1_9",
        "Starts_A2_9",
        "Starts_A3_9",
        "Energy_kWh",
    }
    extras = [c for c in df.columns if c not in expected]
    if extras:
        print(f"  Dropping {len(extras)} unrecognised column(s): {extras}")
        df.drop(columns=extras, inplace=True)

    df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y", errors="coerce")
    df.dropna(subset=["Date"], inplace=True)
    df.sort_values("Date", inplace=True)
    df.reset_index(drop=True, inplace=True)

    numeric_cols = list(expected - {"Date"})
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Year"] = df["Date"].dt.year.astype(int)
    df["Month"] = df["Date"].dt.month.astype(int)
    df["YearMonth"] = df["Date"].dt.to_period("M").astype(str)

    return df


def add_derived_columns(df):
    """Volumes, energy intensity, and per-equipment metrics aligned with the P&ID."""

    fa = df["Flow_Agios_m3"]
    fb = df["Flow_Borehole_m3"]
    fd = df["Flow_Distribution_m3"]

    df["Total_Flow_Delivered_m3"] = fd
    df["Total_Flow_Inflows_To_Tank_m3"] = fa + fb
    df["Total_Flow_Meter_Lines_Sum_m3"] = fa + fb + fd
    df["Flow_Pumped_On_Site_m3"] = fb + fd

    h1, h2, h3 = df["Hours_A1_9"], df["Hours_A2_9"], df["Hours_A3_9"]
    s1, s2, s3 = df["Starts_A1_9"], df["Starts_A2_9"], df["Starts_A3_9"]

    df["Hours_Distribution_Total"] = h1 + h2
    df["Starts_Distribution_Total"] = s1 + s2
    df["Total_Hours"] = h1 + h2 + h3
    df["Total_Starts"] = s1 + s2 + s3

    # Primary KPI: kWh per m³ sent toward D.D. Kostakion (FIT 3.9)
    df["Specific_Energy"] = np.where(
        df["Total_Flow_Delivered_m3"] > 0,
        df["Energy_kWh"] / df["Total_Flow_Delivered_m3"],
        np.nan,
    )
    # Excludes Agios feed (pumped upstream / not at this station)
    df["Specific_Energy_Pumped_On_Site"] = np.where(
        df["Flow_Pumped_On_Site_m3"] > 0,
        df["Energy_kWh"] / df["Flow_Pumped_On_Site_m3"],
        np.nan,
    )

    # Mechanical pumps A1, A2, A3 (alerts & load balance)
    for tag, h, s in (
        ("A1", h1, s1),
        ("A2", h2, s2),
        ("A3", h3, s3),
    ):
        df[f"Utilization_{tag}_pct"] = (h / 24.0) * 100
        df[f"StartsPerHour_{tag}"] = np.where(h > 0, s / h, np.nan)

    meter_sum = df["Total_Flow_Meter_Lines_Sum_m3"]
    df["FlowShare_Line_Agios_pct"] = np.where(meter_sum > 0, (fa / meter_sum) * 100, np.nan)
    df["FlowShare_Line_Borehole_pct"] = np.where(meter_sum > 0, (fb / meter_sum) * 100, np.nan)
    df["FlowShare_Line_Distribution_pct"] = np.where(meter_sum > 0, (fd / meter_sum) * 100, np.nan)

    # Borehole: FIT 2.9 / A3.9 hours
    df["FlowRate_Borehole_m3h"] = np.where(h3 > 0, fb / h3, np.nan)
    # Distribution: one meter for two pumps; sum of run-hours approximates duty/standby sequencing
    h_dist = df["Hours_Distribution_Total"]
    df["FlowRate_Distribution_m3h"] = np.where(h_dist > 0, fd / h_dist, np.nan)
    df["Utilization_Distribution_Dominant_pct"] = (np.maximum(h1, h2) / 24.0) * 100
    df["StartsPerHour_Distribution"] = np.where(
        h_dist > 0, df["Starts_Distribution_Total"] / h_dist, np.nan
    )

    # Inlet: no local pump — average continuous equivalent rate over the day
    df["FlowRate_Agios_daily_avg_m3h"] = np.where(fa > 0, fa / 24.0, np.nan)

    df["Pumps_Active"] = (
        (h1 > 0).astype(int) + (h2 > 0).astype(int) + (h3 > 0).astype(int)
    )

    return df


def safe(val):
    if val is None:
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, pd.Period):
        return str(val)
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating, float)):
        if np.isnan(val) or np.isinf(val):
            return None
        return round(float(val), 4)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    return val


def overview_stats(df):
    n = len(df)
    if n == 0:
        return {}

    delivered = df["Total_Flow_Delivered_m3"].sum()
    inflows = df["Total_Flow_Inflows_To_Tank_m3"].sum()
    meter_sum = df["Total_Flow_Meter_Lines_Sum_m3"].sum()
    pumped_site = df["Flow_Pumped_On_Site_m3"].sum()
    total_energy = df["Energy_kWh"].sum()
    total_hours = df["Total_Hours"].sum()
    total_starts = df["Total_Starts"].sum()
    peak_idx = df["Total_Flow_Delivered_m3"].idxmax()

    return {
        "total_flow_m3": safe(delivered),
        "total_flow_inflows_to_tank_m3": safe(inflows),
        "total_flow_meter_lines_sum_m3": safe(meter_sum),
        "total_flow_pumped_on_site_m3": safe(pumped_site),
        "total_energy_kwh": safe(total_energy),
        "total_hours": safe(total_hours),
        "total_starts": safe(int(total_starts)),
        "specific_energy_kwh_per_m3": safe(total_energy / delivered) if delivered > 0 else None,
        "specific_energy_kwh_per_m3_pumped_on_site": safe(total_energy / pumped_site)
        if pumped_site > 0
        else None,
        "avg_daily_flow_m3": safe(delivered / n),
        "avg_daily_energy_kwh": safe(total_energy / n),
        "avg_daily_hours": safe(total_hours / n),
        "peak_day": {
            "date": safe(df.loc[peak_idx, "Date"]),
            "flow_m3": safe(df.loc[peak_idx, "Total_Flow_Delivered_m3"]),
        },
        "lowest_day": {
            "date": safe(df.loc[df["Total_Flow_Delivered_m3"].idxmin(), "Date"]),
            "flow_m3": safe(df["Total_Flow_Delivered_m3"].min()),
        },
        "total_days": n,
        "date_range": {
            "start": safe(df["Date"].min()),
            "end": safe(df["Date"].max()),
        },
    }


def pump_stats(df):
    """Three cards: inlet line, borehole line+pump A3, distribution line+pumps A1&A2."""
    meter_sum = df["Total_Flow_Meter_Lines_Sum_m3"].sum()
    n = len(df)
    fa = df["Flow_Agios_m3"]
    fb = df["Flow_Borehole_m3"]
    fd = df["Flow_Distribution_m3"]
    h1, h2, h3 = df["Hours_A1_9"], df["Hours_A2_9"], df["Hours_A3_9"]
    s1, s2, s3 = df["Starts_A1_9"], df["Starts_A2_9"], df["Starts_A3_9"]

    running_b = df[h3 > 0]
    nrb = len(running_b)
    h_dist = df["Hours_Distribution_Total"]
    running_d = df[h_dist > 0]
    nrd = len(running_d)

    return {
        "pump_1": {
            "role": "inlet_external_feed",
            "has_local_pump": False,
            "equipment_note_el": "Η ενέργεια αντλίας για την παροχή από Άγιο Γεωργίο δεν μετράται στον σταθμό.",
            "total_flow_m3": safe(fa.sum()),
            "flow_share_pct": safe((fa.sum() / meter_sum) * 100) if meter_sum > 0 else None,
            "total_hours": None,
            "total_starts": None,
            "avg_flow_rate_m3h": safe(df["FlowRate_Agios_daily_avg_m3h"].mean()),
            "avg_utilization_pct": None,
            "avg_starts_per_hour": None,
            "days_active": int((fa > 0).sum()),
            "days_inactive": int((fa <= 0).sum()),
            "max_daily_hours": None,
            "max_daily_flow_m3": safe(fa.max()),
        },
        "pump_2": {
            "role": "borehole",
            "has_local_pump": True,
            "pump_id": "A3.9",
            "flow_meter": "FIT 2.9",
            "total_flow_m3": safe(fb.sum()),
            "flow_share_pct": safe((fb.sum() / meter_sum) * 100) if meter_sum > 0 else None,
            "total_hours": safe(h3.sum()),
            "total_starts": safe(int(s3.sum())),
            "avg_flow_rate_m3h": safe(running_b["FlowRate_Borehole_m3h"].mean()) if nrb > 0 else None,
            "avg_utilization_pct": safe(df["Utilization_A3_pct"].mean()) if n > 0 else None,
            "avg_starts_per_hour": safe(running_b["StartsPerHour_A3"].mean()) if nrb > 0 else None,
            "days_active": nrb,
            "days_inactive": n - nrb,
            "max_daily_hours": safe(h3.max()),
            "max_daily_flow_m3": safe(fb.max()),
        },
        "pump_3": {
            "role": "distribution",
            "has_local_pump": True,
            "pump_ids": ["A1.9", "A2.9"],
            "flow_meter": "FIT 3.9",
            "total_flow_m3": safe(fd.sum()),
            "flow_share_pct": safe((fd.sum() / meter_sum) * 100) if meter_sum > 0 else None,
            "total_hours": safe(h_dist.sum()),
            "total_starts": safe(int(df["Starts_Distribution_Total"].sum())),
            "avg_flow_rate_m3h": safe(running_d["FlowRate_Distribution_m3h"].mean()) if nrd > 0 else None,
            "avg_utilization_pct": safe(df["Utilization_Distribution_Dominant_pct"].mean())
            if n > 0
            else None,
            "avg_starts_per_hour": safe(running_d["StartsPerHour_Distribution"].mean())
            if nrd > 0
            else None,
            "days_active": nrd,
            "days_inactive": n - nrd,
            "max_daily_hours": safe(np.maximum(h1, h2).max()),
            "max_daily_flow_m3": safe(fd.max()),
        },
    }


def _agg_dict():
    return {
        "Total_Flow_Delivered_m3": ("Total_Flow_Delivered_m3", "sum"),
        "Total_Flow_Inflows_To_Tank_m3": ("Total_Flow_Inflows_To_Tank_m3", "sum"),
        "Total_Flow_Meter_Lines_Sum_m3": ("Total_Flow_Meter_Lines_Sum_m3", "sum"),
        "Flow_Pumped_On_Site_m3": ("Flow_Pumped_On_Site_m3", "sum"),
        "Energy_kWh": ("Energy_kWh", "sum"),
        "Total_Hours": ("Total_Hours", "sum"),
        "Total_Starts": ("Total_Starts", "sum"),
        "Flow_Agios_m3": ("Flow_Agios_m3", "sum"),
        "Flow_Borehole_m3": ("Flow_Borehole_m3", "sum"),
        "Flow_Distribution_m3": ("Flow_Distribution_m3", "sum"),
        "Hours_A1_9": ("Hours_A1_9", "sum"),
        "Hours_A2_9": ("Hours_A2_9", "sum"),
        "Hours_A3_9": ("Hours_A3_9", "sum"),
        "Starts_A1_9": ("Starts_A1_9", "sum"),
        "Starts_A2_9": ("Starts_A2_9", "sum"),
        "Starts_A3_9": ("Starts_A3_9", "sum"),
        "Pumps_Active_Avg": ("Pumps_Active", "mean"),
        "days": ("Date", "count"),
    }


def _add_specific_energy(agg):
    d = agg["Total_Flow_Delivered_m3"]
    ps = agg["Flow_Pumped_On_Site_m3"]
    agg["Specific_Energy"] = np.where(d > 0, agg["Energy_kWh"] / d, np.nan)
    agg["Specific_Energy_Pumped_On_Site"] = np.where(
        ps > 0, agg["Energy_kWh"] / ps, np.nan
    )
    agg["Avg_Daily_Flow"] = agg["Total_Flow_Delivered_m3"] / agg["days"]
    return agg


def _row_to_record(row, period_key, period_val):
    return {
        period_key: str(period_val),
        "total_flow_m3": safe(row["Total_Flow_Delivered_m3"]),
        "total_flow_inflows_to_tank_m3": safe(row["Total_Flow_Inflows_To_Tank_m3"]),
        "total_flow_meter_lines_sum_m3": safe(row["Total_Flow_Meter_Lines_Sum_m3"]),
        "total_flow_pumped_on_site_m3": safe(row["Flow_Pumped_On_Site_m3"]),
        "energy_kwh": safe(row["Energy_kWh"]),
        "specific_energy": safe(row["Specific_Energy"]),
        "specific_energy_pumped_on_site": safe(row["Specific_Energy_Pumped_On_Site"]),
        "total_hours": safe(row["Total_Hours"]),
        "total_starts": safe(row["Total_Starts"]),
        "flow_agios_m3": safe(row["Flow_Agios_m3"]),
        "flow_borehole_m3": safe(row["Flow_Borehole_m3"]),
        "flow_distribution_m3": safe(row["Flow_Distribution_m3"]),
        "hours_a1_9": safe(row["Hours_A1_9"]),
        "hours_a2_9": safe(row["Hours_A2_9"]),
        "hours_a3_9": safe(row["Hours_A3_9"]),
        "starts_a1_9": safe(row["Starts_A1_9"]),
        "starts_a2_9": safe(row["Starts_A2_9"]),
        "starts_a3_9": safe(row["Starts_A3_9"]),
        "avg_daily_flow": safe(row["Avg_Daily_Flow"]) if "Avg_Daily_Flow" in row.index else None,
        "avg_pumps_active": safe(row["Pumps_Active_Avg"]) if "Pumps_Active_Avg" in row.index else None,
        "days": safe(row["days"]) if "days" in row.index else None,
    }


def build_daily_series(df):
    cols_out = [
        "Date",
        "Total_Flow_Delivered_m3",
        "Total_Flow_Inflows_To_Tank_m3",
        "Total_Flow_Meter_Lines_Sum_m3",
        "Flow_Pumped_On_Site_m3",
        "Energy_kWh",
        "Specific_Energy",
        "Specific_Energy_Pumped_On_Site",
        "Total_Hours",
        "Total_Starts",
        "Pumps_Active",
        "Flow_Agios_m3",
        "Flow_Borehole_m3",
        "Flow_Distribution_m3",
        "Hours_A1_9",
        "Hours_A2_9",
        "Hours_A3_9",
        "Starts_A1_9",
        "Starts_A2_9",
        "Starts_A3_9",
        "FlowRate_Borehole_m3h",
        "FlowRate_Distribution_m3h",
        "FlowRate_Agios_daily_avg_m3h",
        "Utilization_A1_pct",
        "Utilization_A2_pct",
        "Utilization_A3_pct",
        "Utilization_Distribution_Dominant_pct",
    ]
    subset = df[cols_out].copy()
    subset["Date"] = subset["Date"].dt.strftime("%Y-%m-%d")
    subset.rename(columns={"Date": "date"}, inplace=True)
    # Single name for charts (delivered volume)
    subset.rename(columns={"Total_Flow_Delivered_m3": "total_flow_m3"}, inplace=True)

    records = subset.to_dict(orient="records")
    for rec in records:
        for k, v in rec.items():
            rec[k] = safe(v)
    return records


def build_monthly_series(df):
    agg = df.groupby("YearMonth").agg(**_agg_dict()).reset_index()
    agg = _add_specific_energy(agg)
    return [_row_to_record(row, "period", row["YearMonth"]) for _, row in agg.iterrows()]


def build_yearly_series(df):
    agg = df.groupby("Year").agg(**_agg_dict()).reset_index()
    agg = _add_specific_energy(agg)
    return [_row_to_record(row, "period", int(row["Year"])) for _, row in agg.iterrows()]


def detect_alerts(df):
    short_cycling = []
    for pump_label, col in (
        ("A1.9", "StartsPerHour_A1"),
        ("A2.9", "StartsPerHour_A2"),
        ("A3.9", "StartsPerHour_A3"),
    ):
        flagged = df[df[col] > SHORT_CYCLING_THRESHOLD]
        for _, row in flagged.iterrows():
            hc = {
                "A1.9": "Hours_A1_9",
                "A2.9": "Hours_A2_9",
                "A3.9": "Hours_A3_9",
            }[pump_label]
            scol = {
                "A1.9": "Starts_A1_9",
                "A2.9": "Starts_A2_9",
                "A3.9": "Starts_A3_9",
            }[pump_label]
            short_cycling.append(
                {
                    "date": safe(row["Date"]),
                    "pump": pump_label,
                    "starts_per_hour": safe(row[col]),
                    "starts": safe(row[scol]),
                    "hours": safe(row[hc]),
                }
            )
    short_cycling.sort(key=lambda x: x["date"])

    downtime = df[df["Pumps_Active"] == 0]
    downtime_days = [{"date": safe(row["Date"])} for _, row in downtime.iterrows()]

    high_util = []
    for pump_label, hc in (
        ("A1.9", "Hours_A1_9"),
        ("A2.9", "Hours_A2_9"),
        ("A3.9", "Hours_A3_9"),
    ):
        umap = {"A1.9": "Utilization_A1_pct", "A2.9": "Utilization_A2_pct", "A3.9": "Utilization_A3_pct"}
        flagged = df[df[hc] > HIGH_UTIL_HOURS]
        for _, row in flagged.iterrows():
            high_util.append(
                {
                    "date": safe(row["Date"]),
                    "pump": pump_label,
                    "hours": safe(row[hc]),
                    "utilization_pct": safe(row[umap[pump_label]]),
                }
            )
    high_util.sort(key=lambda x: x["date"])

    se = df["Specific_Energy"].dropna()
    if len(se) > 0:
        q95 = se.quantile(0.95)
        low_efficiency = df[df["Specific_Energy"] > q95].copy()
    else:
        low_efficiency = df.iloc[0:0].copy()
    low_eff_days = []
    for _, row in low_efficiency.iterrows():
        low_eff_days.append(
            {
                "date": safe(row["Date"]),
                "specific_energy": safe(row["Specific_Energy"]),
                "total_flow_delivered_m3": safe(row["Total_Flow_Delivered_m3"]),
                "energy_kwh": safe(row["Energy_kWh"]),
            }
        )

    return {
        "short_cycling_events": short_cycling,
        "downtime_days": downtime_days,
        "high_utilization_days": high_util,
        "low_efficiency_days": low_eff_days,
        "summary": {
            "total_short_cycling_events": len(short_cycling),
            "total_downtime_days": len(downtime_days),
            "total_high_utilization_days": len(high_util),
            "total_low_efficiency_days": len(low_eff_days),
        },
    }


def compute_comparisons(df):
    yearly = (
        df.groupby("Year")
        .agg(
            total_flow=("Total_Flow_Delivered_m3", "sum"),
            total_energy=("Energy_kWh", "sum"),
            total_hours=("Total_Hours", "sum"),
            total_starts=("Total_Starts", "sum"),
            days=("Date", "count"),
        )
        .reset_index()
    )

    yoy = {}
    for idx in range(1, len(yearly)):
        year = int(yearly.iloc[idx]["Year"])
        prev, curr = yearly.iloc[idx - 1], yearly.iloc[idx]

        def pct_change(curr_v, prev_v):
            return safe(((curr_v - prev_v) / prev_v) * 100) if prev_v > 0 else None

        yoy[str(year)] = {
            "flow_change_pct": pct_change(curr["total_flow"], prev["total_flow"]),
            "energy_change_pct": pct_change(curr["total_energy"], prev["total_energy"]),
            "hours_change_pct": pct_change(curr["total_hours"], prev["total_hours"]),
            "starts_change_pct": pct_change(curr["total_starts"], prev["total_starts"]),
        }

    seasonal = df.groupby("Month").agg(
        avg_daily_flow=("Total_Flow_Delivered_m3", "mean"),
        avg_daily_energy=("Energy_kWh", "mean"),
        avg_daily_hours=("Total_Hours", "mean"),
        total_flow=("Total_Flow_Delivered_m3", "sum"),
        total_energy=("Energy_kWh", "sum"),
        days_sampled=("Date", "count"),
    ).reset_index()

    seasonal_dict = {}
    for _, row in seasonal.iterrows():
        seasonal_dict[str(int(row["Month"]))] = {
            "avg_daily_flow_m3": safe(row["avg_daily_flow"]),
            "avg_daily_energy_kwh": safe(row["avg_daily_energy"]),
            "avg_daily_hours": safe(row["avg_daily_hours"]),
            "total_flow_m3": safe(row["total_flow"]),
            "total_energy_kwh": safe(row["total_energy"]),
            "days_sampled": safe(row["days_sampled"]),
        }

    pump_yearly = {
        "pump_1": [
            {
                "year": str(int(row["Year"])),
                "flow_m3": safe(row["Flow_Agios_m3"]),
                "hours": None,
                "starts": None,
                "flow_rate_m3h": None,
            }
            for _, row in df.groupby("Year")
            .agg(Flow_Agios_m3=("Flow_Agios_m3", "sum"))
            .reset_index()
            .iterrows()
        ],
        "pump_2": [
            {
                "year": str(int(row["Year"])),
                "flow_m3": safe(row["Flow_Borehole_m3"]),
                "hours": safe(row["Hours_A3_9"]),
                "starts": safe(int(row["Starts_A3_9"])),
                "flow_rate_m3h": safe(row["Flow_Borehole_m3"] / row["Hours_A3_9"])
                if row["Hours_A3_9"] > 0
                else None,
            }
            for _, row in df.groupby("Year")
            .agg(
                Flow_Borehole_m3=("Flow_Borehole_m3", "sum"),
                Hours_A3_9=("Hours_A3_9", "sum"),
                Starts_A3_9=("Starts_A3_9", "sum"),
            )
            .reset_index()
            .iterrows()
        ],
        "pump_3": [
            {
                "year": str(int(row["Year"])),
                "flow_m3": safe(row["Flow_Distribution_m3"]),
                "hours": safe(row["Hours_Dist"]),
                "starts": safe(int(row["Starts_Dist"])),
                "flow_rate_m3h": safe(row["Flow_Distribution_m3"] / row["Hours_Dist"])
                if row["Hours_Dist"] > 0
                else None,
            }
            for _, row in df.groupby("Year")
            .agg(
                Flow_Distribution_m3=("Flow_Distribution_m3", "sum"),
                Hours_Dist=("Hours_Distribution_Total", "sum"),
                Starts_Dist=("Starts_Distribution_Total", "sum"),
            )
            .reset_index()
            .iterrows()
        ],
    }

    return {
        "year_over_year": yoy,
        "seasonal_averages": seasonal_dict,
        "pump_yearly_trend": pump_yearly,
    }


def compute_load_balance(df):
    """Shares of recorded line flows; separate shares of mechanical pump hours."""
    results = {}

    for period_label, group in [("all_time", df)] + list(df.groupby("Year")):
        if isinstance(period_label, (int, np.integer)):
            period_label = str(period_label)

        fa = group["Flow_Agios_m3"].sum()
        fb = group["Flow_Borehole_m3"].sum()
        fd = group["Flow_Distribution_m3"].sum()
        meter = fa + fb + fd

        h1 = group["Hours_A1_9"].sum()
        h2 = group["Hours_A2_9"].sum()
        h3 = group["Hours_A3_9"].sum()
        th = h1 + h2 + h3

        flow_line_shares = (
            [(fa / meter * 100), (fb / meter * 100), (fd / meter * 100)] if meter > 0 else [0, 0, 0]
        )
        pump_hour_shares = (
            [(h1 / th * 100), (h2 / th * 100), (h3 / th * 100)] if th > 0 else [0, 0, 0]
        )

        results[period_label] = {
            "flow_line_shares_pct": [safe(s) for s in flow_line_shares],
            "flow_line_labels": [
                "FIT 1.9 (Άγιος Γεωργίος)",
                "FIT 2.9 (γεώτρηση)",
                "FIT 3.9 (προς δίκτυο)",
            ],
            "mechanical_pump_hours_shares_pct": [safe(s) for s in pump_hour_shares],
            "mechanical_pump_labels": ["Αντλία A1.9", "Αντλία A2.9", "Αντλία A3.9"],
            "max_flow_line_imbalance_pct": safe(max(abs(s - 100 / 3) for s in flow_line_shares)),
            "max_pump_hours_imbalance_pct": safe(max(abs(s - 100 / 3) for s in pump_hour_shares)),
            "dominant_flow_line_index": int(np.argmax([fa, fb, fd])) + 1,
            "dominant_pump_hours_index": int(np.argmax([h1, h2, h3])) + 1,
        }

    return results


def main():
    print("=" * 60)
    print("  AQUA-STAT: Water Pumping Station Data Processor")
    print("=" * 60)

    if not os.path.exists(BASE_DIR):
        print(f"\nFATAL: ANAFORES folder not found at:\n  {BASE_DIR}")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("\n[1/7] Reading XLS files...")
    raw_df = read_all_monthly_files()
    print(f"  Raw rows collected: {len(raw_df)}")

    print("\n[2/7] Cleaning columns and dates...")
    df = clean_columns(raw_df)
    print(f"  Clean rows: {len(df)}")
    print(f"  Date range: {df['Date'].min().date()} -> {df['Date'].max().date()}")
    print(f"  Columns: {list(df.columns)}")

    print("\n[3/7] Computing derived metrics...")
    df = add_derived_columns(df)

    print("\n[4/7] Aggregating overview statistics...")
    ov_all = overview_stats(df)
    ov_year = {str(y): overview_stats(g) for y, g in df.groupby("Year")}
    ov_month = {ym: overview_stats(g) for ym, g in df.groupby("YearMonth")}

    print("\n[5/7] Aggregating per-line / per-pump statistics...")
    pp_all = pump_stats(df)
    pp_year = {str(y): pump_stats(g) for y, g in df.groupby("Year")}
    pp_month = {ym: pump_stats(g) for ym, g in df.groupby("YearMonth")}

    print("\n[6/7] Building time series (daily / monthly / yearly)...")
    ts_daily = build_daily_series(df)
    ts_monthly = build_monthly_series(df)
    ts_yearly = build_yearly_series(df)

    print("\n[7/7] Detecting alerts, comparisons, load balance...")
    alerts = detect_alerts(df)
    comparisons = compute_comparisons(df)
    load_balance = compute_load_balance(df)

    stats = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "station_name": "Δ8 ΤΣΕ 9 – Υ/Π ΚΩΣΤΑΚΙΩΝ",
            "station_id": "TSE9",
            "hydraulic_notes_el": (
                "Το FIT 1.9 μετρά είσοδο από Άγιο Γεωργίο (ενέργεια αλλού). "
                "Το FIT 2.9 με την A3.9 αφορά γεώτρηση. Το FIT 3.9 μετρά συνδυασμένη παροχή των A1.9 και A2.9 προς το δίκτυο. "
                "Η ειδική ενέργεια (kWh/m³) υπολογίζεται ως kWh διαιρούμενα με όγκο FIT 3.9 (νερό προς δίκτυο)."
            ),
            "pumps": [
                {"id": i + 1, "label": LINE_CARD_LABELS[i]} for i in range(3)
            ],
            "lines": [
                {"id": 1, "label": LINE_CARD_LABELS[0], "fit": "FIT 1.9", "pumps_at_station": []},
                {"id": 2, "label": LINE_CARD_LABELS[1], "fit": "FIT 2.9", "pumps_at_station": ["A3.9"]},
                {
                    "id": 3,
                    "label": LINE_CARD_LABELS[2],
                    "fit": "FIT 3.9",
                    "pumps_at_station": ["A1.9", "A2.9"],
                },
            ],
            "total_days": len(df),
            "years_covered": sorted([int(y) for y in df["Year"].unique()]),
        },
        "overview": {
            "all_time": ov_all,
            "per_year": ov_year,
            "per_month": ov_month,
        },
        "per_pump": {
            "all_time": pp_all,
            "per_year": pp_year,
            "per_month": pp_month,
        },
        "time_series": {
            "daily": ts_daily,
            "monthly": ts_monthly,
            "yearly": ts_yearly,
        },
        "alerts": alerts,
        "comparisons": comparisons,
        "load_balance": load_balance,
    }

    csv_path = os.path.join(OUTPUT_DIR, "master_daily.csv")
    json_path = os.path.join(OUTPUT_DIR, "stats.json")
    excel_path = os.path.join(OUTPUT_DIR, "master_daily.xlsx")

    df_export = df.copy()
    df_export["Date"] = df_export["Date"].dt.strftime("%Y-%m-%d")
    df_export.to_csv(csv_path, index=False, encoding="utf-8-sig")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        df_export.to_excel(excel_path, index=False, sheet_name="Daily Data", engine="openpyxl")

        wb = load_workbook(excel_path)
        ws = wb["Daily Data"]

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )

        for col in range(1, len(df_export.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

            col_letter = get_column_letter(col)
            col_name = df_export.columns[col - 1]
            data_max_len = 0
            if not df_export.empty:
                data_max_len = df_export[col_name].astype(str).map(len).max()
            adjusted_width = min(max(data_max_len, len(str(col_name))) + 4, 30)
            ws.column_dimensions[col_letter].width = adjusted_width

        for row in range(2, len(df_export) + 2):
            for col in range(1, len(df_export.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        ws.freeze_panes = "A2"
        wb.save(excel_path)
    except Exception as e:
        print(f"  Warning: Could not save/format Excel file ({e})")

    print("\n" + "=" * 60)
    print("  DONE!")
    print("=" * 60)
    print(f"  CSV   -> {csv_path}")
    print(f"  JSON  -> {json_path}")
    print(f"  Excel -> {excel_path}")
    print(f"  Total days processed : {len(df)}")
    print(f"  Volume to network (FIT 3.9): {ov_all['total_flow_m3']:,.0f} m³")
    print(f"  Total energy consumed     : {ov_all['total_energy_kwh']:,.0f} kWh")
    se = ov_all["specific_energy_kwh_per_m3"]
    print(f"  Specific energy (kWh/m³ delivered): {se:.4f}" if se else "  Specific energy: N/A")
    s = alerts["summary"]
    print(
        f"  Alerts: {s['total_short_cycling_events']} short-cycling, "
        f"{s['total_downtime_days']} downtime, "
        f"{s['total_high_utilization_days']} high-util, "
        f"{s['total_low_efficiency_days']} low-efficiency"
    )
    print("=" * 60)


if __name__ == "__main__":
    main()
